import os
import csv
import json
import io
import datetime
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, make_response, jsonify
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(level=logging.DEBUG)

# Initialize Flask application
app = Flask(__name__)
app.secret_key = os.environ.get("SESSION_SECRET", "healthcare-shift-logger-secret")

# Initialize in-memory storage for shifts
if 'shifts' not in app.config:
    app.config['shifts'] = []

# Helper function to validate shift data
def validate_shift_data(name, role, start_time, end_time):
    errors = []
    
    if not name:
        errors.append("Staff name is required")
    
    if not role:
        errors.append("Staff role is required")
    
    if not start_time:
        errors.append("Shift start time is required")
    
    if not end_time:
        errors.append("Shift end time is required")
    
    # Check if end time is after start time
    if start_time and end_time:
        try:
            start_dt = datetime.datetime.strptime(start_time, '%Y-%m-%dT%H:%M')
            end_dt = datetime.datetime.strptime(end_time, '%Y-%m-%dT%H:%M')
            
            if end_dt <= start_dt:
                errors.append("Shift end time must be after start time")
                
            # Check for duplicate shifts
            for shift in app.config['shifts']:
                shift_start = datetime.datetime.strptime(shift['start_time'], '%Y-%m-%dT%H:%M')
                shift_end = datetime.datetime.strptime(shift['end_time'], '%Y-%m-%dT%H:%M')
                
                # Check for exact duplicates (same name, role, start and end time)
                if (name.lower() == shift['name'].lower() and 
                    role == shift['role'] and 
                    start_dt == shift_start and 
                    end_dt == shift_end):
                    errors.append("This exact shift already exists")
                    break
                
                # Check for overlapping shifts for the same person
                if name.lower() == shift['name'].lower():
                    # Check if the new shift overlaps with an existing shift
                    if (start_dt < shift_end and end_dt > shift_start):
                        errors.append(f"This shift overlaps with an existing shift for {name}")
                        break
                        
        except ValueError:
            errors.append("Invalid time format")
    
    return errors

# Home page route
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Check if this is a filter/sort form submission
        if 'filter_submit' in request.form:
            # Store filter and sort preferences in session
            session['filter_role'] = request.form.get('filter_role', 'all')
            session['sort_by'] = request.form.get('sort_by', 'name')
            session['sort_order'] = request.form.get('sort_order', 'asc')
            return redirect(url_for('index'))
        
        # Otherwise, it's a new shift submission
        name = request.form.get('name', '').strip()
        role = request.form.get('role', '').strip()
        start_time = request.form.get('start_time', '')
        end_time = request.form.get('end_time', '')
        
        # Validate the input
        errors = validate_shift_data(name, role, start_time, end_time)
        
        if errors:
            for error in errors:
                flash(error, 'danger')
        else:
            # Format times for display
            start_dt = datetime.datetime.strptime(start_time, '%Y-%m-%dT%H:%M')
            end_dt = datetime.datetime.strptime(end_time, '%Y-%m-%dT%H:%M')
            
            display_start = start_dt.strftime('%Y-%m-%d %H:%M')
            display_end = end_dt.strftime('%Y-%m-%d %H:%M')
            
            # Create a unique ID for the shift (timestamp)
            shift_id = str(int(datetime.datetime.now().timestamp()))
            
            # Add new shift to memory
            new_shift = {
                'id': shift_id,
                'name': name,
                'role': role,
                'start_time': start_time,
                'end_time': end_time,
                'display_start': display_start,
                'display_end': display_end
            }
            
            app.config['shifts'].append(new_shift)
            flash('Shift added successfully!', 'success')
            
            # Optionally save to JSON file
            save_to_json()
            
            return redirect(url_for('index'))
    
    # Get filter and sort settings from session or use defaults
    filter_role = session.get('filter_role', 'all')
    sort_by = session.get('sort_by', 'name')
    sort_order = session.get('sort_order', 'asc')
    
    # Get all shifts
    all_shifts = app.config['shifts']
    
    # Calculate dashboard statistics
    total_shifts = len(all_shifts)
    
    # Count shifts by role
    shifts_by_role = {}
    for shift in all_shifts:
        role = shift['role']
        if role in shifts_by_role:
            shifts_by_role[role] += 1
        else:
            shifts_by_role[role] = 1
    
    # Calculate upcoming shifts (today and tomorrow)
    today = datetime.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = today + datetime.timedelta(days=1)
    day_after_tomorrow = today + datetime.timedelta(days=2)
    
    upcoming_shifts_today = 0
    upcoming_shifts_tomorrow = 0
    total_shift_duration = 0
    
    for shift in all_shifts:
        try:
            start_dt = datetime.datetime.strptime(shift['start_time'], '%Y-%m-%dT%H:%M')
            end_dt = datetime.datetime.strptime(shift['end_time'], '%Y-%m-%dT%H:%M')
            
            # Check if shift is today or tomorrow
            if today <= start_dt < tomorrow:
                upcoming_shifts_today += 1
            elif tomorrow <= start_dt < day_after_tomorrow:
                upcoming_shifts_tomorrow += 1
            
            # Calculate shift duration in hours
            duration = (end_dt - start_dt).total_seconds() / 3600
            total_shift_duration += duration
        except (ValueError, TypeError):
            # Skip shifts with invalid dates
            pass
    
    # Calculate average shift length
    avg_shift_length = 0
    if total_shifts > 0:
        avg_shift_length = round(total_shift_duration / total_shifts, 1)
    
    # Apply role filter if not "all"
    filtered_shifts = all_shifts
    if filter_role != 'all':
        filtered_shifts = [shift for shift in filtered_shifts if shift['role'] == filter_role]
    
    # Apply sorting
    if sort_by == 'name':
        filtered_shifts = sorted(filtered_shifts, key=lambda x: x['name'].lower(), reverse=(sort_order == 'desc'))
    elif sort_by == 'role':
        filtered_shifts = sorted(filtered_shifts, key=lambda x: x['role'], reverse=(sort_order == 'desc'))
    elif sort_by == 'start_time':
        filtered_shifts = sorted(filtered_shifts, key=lambda x: x['start_time'], reverse=(sort_order == 'desc'))
    
    # Get unique roles for filter dropdown
    all_roles = sorted(set(shift['role'] for shift in all_shifts))
    
    # Render template with filtered and sorted shifts plus dashboard data
    return render_template('index.html', 
                          shifts=filtered_shifts, 
                          filter_role=filter_role,
                          sort_by=sort_by,
                          sort_order=sort_order,
                          all_roles=all_roles,
                          now=datetime.datetime.now(),
                          # Dashboard data
                          total_shifts=total_shifts,
                          shifts_by_role=shifts_by_role,
                          upcoming_shifts_today=upcoming_shifts_today,
                          upcoming_shifts_tomorrow=upcoming_shifts_tomorrow,
                          avg_shift_length=avg_shift_length)

# Delete shift route
@app.route('/delete/<shift_id>', methods=['POST'])
def delete_shift(shift_id):
    shifts = app.config['shifts']
    app.config['shifts'] = [shift for shift in shifts if shift['id'] != shift_id]
    
    # Optionally save to JSON file
    save_to_json()
    
    flash('Shift deleted successfully!', 'success')
    return redirect(url_for('index'))

# Export to CSV route
@app.route('/export-csv')
def export_csv():
    # Create a CSV string
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['Name', 'Role', 'Shift Start', 'Shift End'])
    
    # Write data
    for shift in app.config['shifts']:
        writer.writerow([
            shift['name'], 
            shift['role'], 
            shift['display_start'], 
            shift['display_end']
        ])
    
    # Create response
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=shifts.csv"
    response.headers["Content-type"] = "text/csv"
    
    return response

# Backup/Download JSON route
@app.route('/backup-json')
def backup_json():
    try:
        if os.path.exists('shifts.json'):
            return send_file(
                'shifts.json',
                mimetype='application/json',
                as_attachment=True,
                download_name='shifts_backup.json'
            )
        else:
            # If the file doesn't exist yet, create it first
            save_to_json()
            return send_file(
                'shifts.json',
                mimetype='application/json',
                as_attachment=True,
                download_name='shifts_backup.json'
            )
    except Exception as e:
        logging.error(f"Error creating backup: {e}")
        flash('Could not create backup file', 'danger')
        return redirect(url_for('index'))

# Restore from JSON backup
@app.route('/restore-json', methods=['POST'])
def restore_json():
    if 'file' not in request.files:
        flash('No file selected', 'danger')
        return redirect(url_for('index'))
    
    file = request.files['file']
    
    if file.filename == '':
        flash('No file selected', 'danger')
        return redirect(url_for('index'))
    
    if not file.filename.endswith('.json'):
        flash('Only .json files are allowed', 'danger')
        return redirect(url_for('index'))
    
    try:
        # Read the uploaded JSON file
        json_data = json.loads(file.read().decode('utf-8'))
        
        # Validate the data structure
        if not isinstance(json_data, list):
            flash('Invalid backup file format', 'danger')
            return redirect(url_for('index'))
        
        # Check for required fields in each shift
        required_fields = ['id', 'name', 'role', 'start_time', 'end_time', 'display_start', 'display_end']
        for shift in json_data:
            if not all(field in shift for field in required_fields):
                flash('Invalid backup file structure', 'danger')
                return redirect(url_for('index'))
        
        # Update the application data
        app.config['shifts'] = json_data
        
        # Save to JSON file
        save_to_json()
        
        flash(f'Successfully restored {len(json_data)} shifts from backup', 'success')
    except json.JSONDecodeError:
        flash('Invalid JSON file', 'danger')
    except Exception as e:
        logging.error(f"Error restoring from backup: {e}")
        flash('Error restoring from backup', 'danger')
    
    return redirect(url_for('index'))
        
# Export to Excel (.xlsx) route
@app.route('/export-excel')
def export_excel():
    # Create a workbook and select active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shifts"
    
    # Define header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write header row
    headers = ['Name', 'Role', 'Shift Start', 'Shift End', 'Duration (hours)']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data rows
    row_num = 2
    for shift in app.config['shifts']:
        # Calculate shift duration
        try:
            start_dt = datetime.datetime.strptime(shift['start_time'], '%Y-%m-%dT%H:%M')
            end_dt = datetime.datetime.strptime(shift['end_time'], '%Y-%m-%dT%H:%M')
            duration = round((end_dt - start_dt).total_seconds() / 3600, 1)
        except:
            duration = 0
        
        ws.cell(row=row_num, column=1, value=shift['name'])
        ws.cell(row=row_num, column=2, value=shift['role'])
        ws.cell(row=row_num, column=3, value=shift['display_start'])
        ws.cell(row=row_num, column=4, value=shift['display_end'])
        ws.cell(row=row_num, column=5, value=duration)
        
        row_num += 1
    
    # Add a summary section
    row_num += 2  # Skip a row
    
    # Add summary header
    summary_row = row_num
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
    row_num += 1
    
    # Add total shifts
    ws.cell(row=row_num, column=1, value="Total Shifts:")
    ws.cell(row=row_num, column=2, value=len(app.config['shifts']))
    row_num += 1
    
    # Add shifts by role
    ws.cell(row=row_num, column=1, value="Shifts by Role:").font = Font(bold=True)
    row_num += 1
    
    shifts_by_role = {}
    for shift in app.config['shifts']:
        role = shift['role']
        if role in shifts_by_role:
            shifts_by_role[role] += 1
        else:
            shifts_by_role[role] = 1
    
    for role, count in shifts_by_role.items():
        ws.cell(row=row_num, column=1, value=role)
        ws.cell(row=row_num, column=2, value=count)
        row_num += 1
    
    # Add average shift length
    total_duration = 0
    valid_shifts = 0
    for shift in app.config['shifts']:
        try:
            start_dt = datetime.datetime.strptime(shift['start_time'], '%Y-%m-%dT%H:%M')
            end_dt = datetime.datetime.strptime(shift['end_time'], '%Y-%m-%dT%H:%M')
            duration = (end_dt - start_dt).total_seconds() / 3600
            total_duration += duration
            valid_shifts += 1
        except:
            pass
    
    avg_shift_length = 0
    if valid_shifts > 0:
        avg_shift_length = round(total_duration / valid_shifts, 1)
    
    row_num += 1
    ws.cell(row=row_num, column=1, value="Average Shift Length:")
    ws.cell(row=row_num, column=2, value=f"{avg_shift_length} hours")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="healthcare_shifts.xlsx"
    )

# Helper function to save data to JSON file
def save_to_json():
    try:
        with open('shifts.json', 'w') as f:
            json.dump(app.config['shifts'], f)
    except Exception as e:
        logging.error(f"Error saving to JSON: {e}")
        flash('Could not save data to file', 'warning')

# Helper function to load data from JSON file
def load_from_json():
    try:
        with open('shifts.json', 'r') as f:
            app.config['shifts'] = json.load(f)
    except FileNotFoundError:
        app.config['shifts'] = []
    except Exception as e:
        logging.error(f"Error loading from JSON: {e}")
        app.config['shifts'] = []

# Initialize data at startup
def initialize_data():
    load_from_json()

# Run the application
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
    
# Call initialize_data when app starts
with app.app_context():
    initialize_data()
